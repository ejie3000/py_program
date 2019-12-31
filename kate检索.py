from xml.dom.minidom import parse
import xml.dom.minidom
import pymysql, re, os
import csv
import openpyxl, numpy
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

#=================================================================================================================

class Item:
    #定义基本属性
    itemid = 0
    posx = 0
    posy = 0
    content = ''
    color = 'pink'
    #定义构造方法
    def __init__(self,i,px,py,c):
        self.itemid = i
        self.posx = px
        self.posy = py
        self.content = c


#=================================================================================================================

#=================================================================================================================
#单元格颜色
fill_green = PatternFill("solid", fgColor="2E8B57")  #定义颜色 绿色   已上线
fill_yellow = PatternFill("solid", fgColor="ffff99")  #定义颜色 黄色  制作中
fill_pink = PatternFill("solid", fgColor="ff9999")  #定义颜色 粉色    重复
fill_blue = PatternFill("solid", fgColor="9999ff")  #定义颜色 蓝色    已做待上线
fill_orange = PatternFill("solid", fgColor="ffc000")  #定义颜色 橙色     已寄样未拆
#=================================================================================================================

name = []
xml_file_path = ("C:/Users/3d/Desktop/小黑板服务器/clothdata.xml")
xlsx_file_path = ('C:/Users/3d/Desktop/test/第45周.xlsx')

#===================================================================================================
#取XML中的货号


def xml_num_name():
    DOMTree = xml.dom.minidom.parse(xml_file_path)
    trees = DOMTree.documentElement
    xml_name = trees.getElementsByTagName('clothname')  # 解析结构
    itemnolist = []

    for i in xml_name:

        a = i.childNodes[0].data  #得到货号全名
        b = re.sub('\D', '', a)  #用正则表达式取出字符串中的数字
        itemnolist.append(b)
    return itemnolist


itemnolist = xml_num_name()
print('获取已上线货号')
print(len(itemnolist))


#===================================================================================================
#取文件夹中的货号
def dir_num_name():
    dirs = 'T:/UNIQLO'
    c = os.listdir(dirs)
    dir_name = []
    for d in c:
        dirs_list = re.sub('\D', '', d)
        dir_name.append(dirs_list)
        #print(i)
    return dir_name


dir_name = dir_num_name()
print('获取制作文件夹货号')
print(len(dir_name))


#===================================================================================================
#读取excel中 寄样list 的货号
def load_xlsx():
    wb = openpyxl.load_workbook(xlsx_file_path)  #打开excel文件
    #行 row 列 column 单元格 cell 值 value

    ws = wb['寄样list']  #获取指定工作表
    colC = ws['B']  #获取列
    new_name = []

    for col in colC[1:-1]:
        name.append(col.value)

    for id in name:  #列表去重
        if id not in new_name:
            new_name.append(id)
    return new_name


print('获取' + xlsx_file_path + ' 寄样list货号')
new_name = load_xlsx()

print(len(new_name))

#=================================================================================================================

wb = openpyxl.load_workbook('C:/Users/3d/Desktop/test/第52周.xlsx')  #打开excel文件
#行 row 列 column 单元格 cell 值 value
ws = wb.get_sheet_by_name('素材')
#=================================================================================================================

#=================================================================================================================

len_row = ws.max_row
len_column = ws.max_column

print('------------------------')
print('获取品名数据')
print('总行数', len_row)
print('总列数', len_column)
print('------------------------')

# cell_name = ws.cell(column=7,row=4).value
# cell_num_name = re.sub('\D', '', cell_name)[-6:]  #用正则表达式取出字符串中的数字 取后6位
# print(cell_name)
# print(cell_num_name)

#获取到表格中的数据，实例化,并且加入列表中
itemlist = []
itemchecklist=[]
count = 0
for column_y in range(len_column):
    column_y += 1
    cell_name = ws.cell(column=column_y, row=1).value
    if cell_name == '品名':
        # print('------------------------')
        # print(column_y)
        # print('------------------------')

        for row_x in range(1, len_row):
            row_x += 1
            # print(a,i)
            sheet_name = ws.cell(column=column_y, row=row_x).value
            
            if sheet_name != None :

                sheet_num_name = re.sub('\D', '', sheet_name)[-6:]
                if len(sheet_num_name) >= 6:
                    itemlist.append(Item(count,row_x,column_y,sheet_num_name))
                    #print('{},{},{},{}'.format(itemlist[count].itemid,itemlist[count].posx,itemlist[count].posy,itemlist[count].content))   
                    count += 1
                # print(row_x, column_y)
                # print(sheet_num_name)
                # print('------------------------')


#开始比对数据
for target_list in itemlist:
    if target_list.content not in itemchecklist:
        itemchecklist.append(target_list.content)
       # itemdict.update(target_list.sheet_num_name,target_list)
        #如果在正在做的列表中，颜色设置为黄色
        if (target_list.content in dir_name):
            target_list.color = "yellow"
            ws.cell(target_list.posx, target_list.posy).fill = fill_yellow
        #如果在已上线的列表中，颜色设置为绿色
        elif (target_list.content in itemnolist):
            target_list.color = "green"
            ws.cell(target_list.posx, target_list.posy).fill = fill_green
        #如果在其他表格的列表中，颜色设置为绿色
        elif (target_list.content in new_name):
            target_list.color = "orange"
            ws.cell(target_list.posx, target_list.posy).fill = fill_orange
        #都不存在，则为白色
        else:
            target_list.color = "white"
    else:
        target_list.color = itemlist[target_list.itemid].color
        if(target_list.color == "yellow"):
            ws.cell(target_list.posx, target_list.posy).fill = fill_orange
        elif(target_list.color == "green"):
            ws.cell(target_list.posx, target_list.posy).fill = fill_green
        elif(target_list.color == "orange"):
            ws.cell(target_list.posx, target_list.posy).fill = fill_orange

print('已完成')   


wb.save('52.xlsx')