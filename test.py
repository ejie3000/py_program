import openpyxl,numpy,re
import  pandas  as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
#=================================================================================================================
# xls_read = load_workbook('pyxl_test.xlsx')  # 打开excel文件名为'pyxl_test.xlsx'
# # print(xls_read.sheetnames)  # 查看工作表'pyxl_test.xlsx'中的所有sheet名，以列表形式生成

# print(xls_read.active)  # 查看文件pyxl_test的活动中sheet
# # xls_read.active.title = 'test'  # 将活动中的sheet名称变更为test
# xls_read_sheet = xls_read.active  # 将活动中的sheet赋值给变量

# # xls_read_sheet = xls_read.get_sheet_by_name('test') # 获取excel文件的某一个sheet

# # print(xls_read_sheet['C'])    # 读取sheet中的C列所有数据，数据是以元组形式呈现
# # print(xls_read_sheet['C4'].value) # 读取sheet中'C4'单元格的值
# # print(xls_read_sheet.max_column)  # 查看sheet中最大的列合计值，统计依据是只要单元格含有值，就算一列
# # print(xls_read_sheet.max_row)     # 查看sheet中最大的行合计值，统计依据是只要单元格含有值，就算一行

# # b4 = xls_read_sheet['B4']   # 通过列号+行数 来定位某一个cell
# # print(b4.value)   # 使用.value 来获取某一个cell的值

# # print(xls_read_sheet.cell(column=2,row=4).value)  # 通过某sheet.cell(column=?,row=?).value 来取的某一个单元格的值

# # xls_read_sheet.rows # sheet.rows是一个生成器，把每一行的内容形成一个元组
# # for row in xls_read_sheet.rows:
# #     print(row)
# #     for cell in row:
# #         print(cell.value)
# #

# # for column in xls_read_sheet.columns:   #sheet.columns是一个生成器，遍历每一列 一列的内容形成元组
# #     print(list(column))
# #     print(list(column)[0].value)
# # for i in column:
# #     if i.value:
# #         print(i.value)
# # print(i.value)
#=================================================================================================================

wb = openpyxl.load_workbook('C:/Users/3d/Desktop/test/第45周.xlsx')#打开excel文件
#行 row 列 column 单元格 cell 值 value
ws = wb.get_sheet_by_name('素材')
#=================================================================================================================
#改单元格颜色
# d4 = ws['D4']
fill_green = PatternFill("solid", fgColor="2E8B57")#定义颜色 绿色   已上线
fill_yellow = PatternFill("solid", fgColor="ffff99")#定义颜色 黄色  制作中
fill_pink = PatternFill("solid", fgColor="ff9999")#定义颜色 粉色    重复
fill_blue = PatternFill("solid", fgColor="9999ff")#定义颜色 蓝色    已做待上线
fill_orange = PatternFill("solid", fgColor="ffc000")#定义颜色 橙色     已寄样未拆
# d4.fill = fill_orange
#=================================================================================================================

len_row = ws.max_row
len_column = ws.max_column

print('------------------------')
print('总行数', len_row)
print('总列数', len_column)
print('------------------------')


# cell_name = ws.cell(column=7,row=4).value
# cell_num_name = re.sub('\D', '', cell_name)[-6:]  #用正则表达式取出字符串中的数字 取后6位
# print(cell_name)
# print(cell_num_name)

for column_y in range(len_column):
    column_y+=1
    cell_name = ws.cell(column=column_y,row=1).value
    if cell_name == '品名':
        print('------------------------')
        print(column_y)
        print('------------------------')

        for row_x in range(1,len_row):
            row_x +=1
            # print(a,i)
            sheet_name = ws.cell(column=column_y,row=row_x).value
            
            if sheet_name != None:                
                sheet_num_name = re.sub('\D', '', sheet_name)[-6:]
                print(row_x,column_y)
                print(sheet_num_name)
                print('------------------------')
                ws.cell(column=column_y,row=row_x).fill = fill_orange
    #print(cell_name)

class Item:
    posx =0
    posy =0
    content = ""
    color = "pink"

def __init__(self,posx,posy,content):
    self.posx = posx
    self.posy = posy
    self.content = content


#wb.save('46.xlsx')
